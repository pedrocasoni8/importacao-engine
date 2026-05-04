import pdfplumber
from openpyxl import load_workbook
from openai import OpenAI
import json
import os

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

def processar_pdf(input_pdf, output_excel):

    # =========================
    # 1. EXTRAIR TEXTO DO PDF
    # =========================
    texto = ""
    with pdfplumber.open(input_pdf) as pdf:
        for pagina in pdf.pages:
            texto += pagina.extract_text() or ""

    # =========================
    # 2. IA INTERPRETAR
    # =========================
    prompt = f"""
    Você é um especialista em importação e fechamento financeiro de DI.

    Extraia os seguintes campos do texto abaixo:

    - ICMS
    - AFRMM
    - TOTAL_NF
    - TOTAL_DI
    - VALOR_ANTECIPADO (NOP20)
    - HONORARIOS
    - RETENCOES

    Retorne apenas JSON válido assim:

    {{
      "icms": number,
      "afrmm": number,
      "total_nf": number,
      "total_di": number,
      "antecipado": number,
      "honorarios": number,
      "retencoes": number
    }}

    Texto:
    {texto[:12000]}
    """

    response = client.chat.completions.create(
        model="gpt-4.1-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0
    )

    conteudo = response.choices[0].message.content

    try:
        dados = json.loads(conteudo)
    except:
        dados = {}

    # =========================
    # 3. ABRIR MODELO EXCEL
    # =========================
    wb = load_workbook("modelo.xlsx")  # seu MOD_OVER
    ws = wb.active

    # =========================
    # 4. PREENCHIMENTO (SUAS REGRAS)
    # =========================

    icms = dados.get("icms", 0)
    afrmm = dados.get("afrmm", 0)
    antecipado = dados.get("antecipado", 0)
    honorarios = dados.get("honorarios", 0)
    retencoes = dados.get("retencoes", 0)

    # 👉 ICMS e AFRMM só em OUTROS CUSTOS (ex: O73)
    ws["O73"] = icms + afrmm

    # 👉 TOTAL correto (corrigindo seu erro anterior)
    ws["P17"] = dados.get("total_nf", 0)

    # 👉 SALDO A PAGAR (P69)
    ws["P69"] = dados.get("total_di", 0) - antecipado

    # 👉 NOP50 (HONORÁRIO BRUTO)
    ws["NOP50"] = honorarios

    # 👉 R52 (LÍQUIDO)
    ws["R52"] = honorarios - retencoes

    # =========================
    # 5. SALVAR
    # =========================
    wb.save(output_excel)
